from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 새로운 Sheet 생성
ws.title = "MySheet"
ws.sheet_properties.tabColor = "ff66ff" # RGB 형태로 입력

ws1 = wb.create_sheet("yoursheet") # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) # 2번째 index에 Sheet 생성

new_ws = wb["NewSheet"] # Dict 형태로 Sheet에 접근 가능

print(wb.sheetnames) # 모든 Sheet 이름 확인

# Sheet 복사

new_ws["A1"] = "Text"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

print(wb.sheetnames)

wb.save("sample.xlsx")
wb.close()