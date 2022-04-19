from openpyxl import *
wb = load_workbook("scores.xlsx")
ws = wb.active

# 1
for i in range(0, 10):
    ws["D" + str(i + 2)] = 10

# 2
for i in range(0, 10):
    ws["H" + str(i + 2)] = "=SUM(B" + str(i + 2) + ":G" + str(i + 2) + ")"

wb.save("scores_result.xlsx")

# 3
wb = load_workbook("scores_result.xlsx", data_only=True)
ws = wb.active

for i in range(0, 10):
    if int(ws["H" + str(i + 2)].value) >= 90:
        result = "A"
    elif int(ws["H" + str(i + 2)].value) >= 80 and int(ws["H" + str(i + 2)].value) < 90:
        result = "B"
    elif int(ws["H" + str(i + 2)].value) >= 70 and int(ws["H" + str(i + 2)].value) < 80:
        result = "C"
    elif int(ws["H" + str(i + 2)].value) < 70:
        result = "D"

    if int(ws["B" + str(i + 2)].value) < 5:
        result = "F"

    ws["I" + str(i + 2)] = result

wb.save("scores_result.xlsx")