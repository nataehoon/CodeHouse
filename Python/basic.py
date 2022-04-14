from openpyxl import *
from openpyxl.styles import *
import thickborder

wb = Workbook()
ws = wb.active

# 카테고리 삽입====================================================
ws.append(["No.", "Method", "Kind"])

# 카테고리(row[0]) 글씨스타일 변경
for i in range(1, ws.max_column + 1):
    row = ws.cell(row=1, column=i)
    row.font = Font(bold=True, size=15)

# 데이터 삽입 ======================================================
ws["B2"] = "cv2.IMREAD_(이미지 색상 설정)"
ws["C2"] = "COLOR"
ws["C3"] = "GRAYSCALE"
ws["C4"] = "UNCHANGED"

ws["B5"] = "cv2.COLOR_(회색으로 이미지 변경)"
ws["C5"] = "BGR2GRAY"

ws["B6"] = "cv2.LINE_(선 그릴때 선 스타일)"
ws["C6"] = 4
ws["C7"] = 8
ws["C8"] = "AA"

ws["B9"] = "cv2.FONT_(글꼴 종류)"
ws["C9"] = "HERSHEY_SIMPLEX"
ws["C10"] = "HERSHEY_PLAIN"
ws["C11"] = "HERSHEY_SCRIPT_SIMPLEX"
ws["C12"] = "HERSHEY_TRIPLEX"
ws["C13"] = "ITALC"

ws["B14"] = "cv2.INTER_(보간법)"
ws["C14"] = "AREA"
ws["C15"] = "CUBIC"
ws["C16"] = "LINEAR"

ws["B17"] = "cv2.ROTATE_(회전)"
ws["C17"] = "90_CLOCKWISE"
ws["C18"] = 180
ws["C19"] = "90_COUNTERCLOCKWISE"

ws["B20"] = "cv2.EVENT_(마우스 이벤트)"
ws["C20"] = "LBUTTONDOWN"
ws["C21"] = "LBUTTONDBLCLK"
ws["C22"] = "MOUSEWHEEL"
ws["C23"] = "MOUSEMOVE"

# 셀 넓이 조절
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 25

# 모든 셀 테두리 삽입 및 가운데 정렬==================================
thickborder.border(ws, "A1:C1")
thickborder.border(ws, "A2:C5")
thickborder.border(ws, "A6:C8")
thickborder.border(ws, "A9:C13")
thickborder.border(ws, "A14:C16")
thickborder.border(ws, "A17:C19")
thickborder.border(ws, "A20:C23")

# 가운데 정렬
for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")

# 같은 메소드 병합==================================================
# A컬럼 머지
ws.merge_cells("A2:A4")
ws.merge_cells("A6:A8")
ws.merge_cells("A9:A13")
ws.merge_cells("A14:A16")
ws.merge_cells("A17:A19")
ws.merge_cells("A20:A23")

# B컬럼 머지
ws.merge_cells("B2:B4")
ws.merge_cells("B6:B8")
ws.merge_cells("B9:B13")
ws.merge_cells("B14:B16")
ws.merge_cells("B17:B19")
ws.merge_cells("B20:B23")

# 순번 삽입 및 셀 조절===============================================
# No.
i = 1
for row in range(1, ws.max_row):
    row_status = ws["A{}".format(row +1)]
    if "<MergedCell 'Sheet'" in str(row_status):
        continue
    row_status.value = i
    i += 1

ws["b23"].border = Border(bottom=Side(style="thick"), left=Side(style="thin"), right=Side(style="thin"))
ws["A23"].border = Border(bottom=Side(style="thick"), left=Side(style="thick"), right=Side(style="thin"))

ws.freeze_panes = "B2"
wb.save("Programing.xlsx")