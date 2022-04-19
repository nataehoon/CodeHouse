from openpyxl import *
from openpyxl.styles import *
import thickborder

wb = Workbook()
ws = wb.active

# 카테고리 삽입====================================================
ws.append(["No.", "Method", "Kind"])

# 카테고리(row[0]) 글씨스타일 변경
for i in range(1, ws.max_column + 1):
    row = ws.cell(row=1, column=i)
    row.font = Font(bold=True, size=15)

# 데이터 삽입 ======================================================
num = ws.max_row +1
ws[f"B{num}"] = "cv2.IMREAD_(이미지 색상 설정)"
ws[f"C{num}"] = "COLOR"
ws[f"C{num+1}"] = "GRAYSCALE"
ws[f"C{num+2}"] = "UNCHANGED"
last_cell = num+2
thickborder.border(ws, f"A{num}:C{last_cell}")
ws.merge_cells(f"A{num}:A{last_cell}")
ws.merge_cells(f"B{num}:B{last_cell}")

num = ws.max_row +1
ws[f"B{num}"] = "cv2.COLOR_(회색으로 이미지 변경)"
ws[f"C{num}"] = "BGR2GRAY"
last_cell = num
thickborder.border(ws, f"A{num}:C{last_cell}")
ws.merge_cells(f"A{num}:A{last_cell}")
ws.merge_cells(f"B{num}:B{last_cell}")

num = ws.max_row +1
ws[f"B{num}"] = "cv2.LINE_(선 그릴때 선 스타일)"
ws[f"C{num}"] = 4
ws[f"C{num+1}"] = 8
ws[f"C{num+2}"] = "AA"
last_cell = num+2
thickborder.border(ws, f"A{num}:C{last_cell}")
ws.merge_cells(f"A{num}:A{last_cell}")
ws.merge_cells(f"B{num}:B{last_cell}")

num = ws.max_row +1
ws[f"B{num}"] = "cv2.FONT_(글꼴 종류)"
ws[f"C{num}"] = "HERSHEY_SIMPLEX"
ws[f"C{num+1}"] = "HERSHEY_PLAIN"
ws[f"C{num+2}"] = "HERSHEY_SCRIPT_SIMPLEX"
ws[f"C{num+3}"] = "HERSHEY_TRIPLEX"
ws[f"C{num+4}"] = "ITALC"
last_cell = num+4
thickborder.border(ws, f"A{num}:C{last_cell}")
ws.merge_cells(f"A{num}:A{last_cell}")
ws.merge_cells(f"B{num}:B{last_cell}")

num = ws.max_row +1
ws[f"B{num}"] = "cv2.INTER_(보간법)"
ws[f"C{num}"] = "AREA"
ws[f"C{num+1}"] = "CUBIC"
ws[f"C{num+2}"] = "LINEAR"
last_cell = num+2
thickborder.border(ws, f"A{num}:C{last_cell}")
ws.merge_cells(f"A{num}:A{last_cell}")
ws.merge_cells(f"B{num}:B{last_cell}")

num = ws.max_row +1
ws[f"B{num}"] = "cv2.ROTATE_(회전)"
ws[f"C{num}"] = "90_CLOCKWISE"
ws[f"C{num+1}"] = 180
ws[f"C{num+2}"] = "90_COUNTERCLOCKWISE"
last_cell = num+2
thickborder.border(ws, f"A{num}:C{last_cell}")
ws.merge_cells(f"A{num}:A{last_cell}")
ws.merge_cells(f"B{num}:B{last_cell}")

num = ws.max_row +1
ws[f"B{num}"] = "cv2.EVENT_(마우스 이벤트)"
ws[f"C{num}"] = "LBUTTONDOWN"
ws[f"C{num+1}"] = "LBUTTONDBLCLK"
ws[f"C{num+2}"] = "MOUSEWHEEL"
ws[f"C{num+3}"] = "MOUSEMOVE"
last_cell = num+3
thickborder.border(ws, f"A{num}:C{last_cell}")
ws.merge_cells(f"A{num}:A{last_cell}")
ws.merge_cells(f"B{num}:B{last_cell}")

num = ws.max_row +1
ws[f"B{num}"] = "cv2.THRESH_(검/흰 만 표현)"
ws[f"C{num}"] = "BINARY"
ws[f"C{num+1}"] = "BINARY_INV"
ws[f"C{num+2}"] = "OTSU"
ws[f"C{num+3}"] = "ADAPTIVE_....MEAN"
last_cell = num+3
thickborder.border(ws, f"A{num}:C{last_cell}")
ws.merge_cells(f"A{num}:A{last_cell}")
ws.merge_cells(f"B{num}:B{last_cell}")

num = ws.max_row +1
ws[f"B{num}"] = "cv2.RECT_(사각형으로 인식)"
ws[f"C{num}"] = "EXTERNAL"
ws[f"C{num+1}"] = "LIST"
ws[f"C{num+2}"] = "TREE"
last_cell = num+2
thickborder.border(ws, f"A{num}:C{last_cell}")
ws.merge_cells(f"A{num}:A{last_cell}")
ws.merge_cells(f"B{num}:B{last_cell}")

# 셀 넓이 조절
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 25

# 가운데 정렬
for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")

# 순번 삽입 및 셀 조절===============================================
# No.
i = 1
for row in range(1, ws.max_row):
    row_status = ws["A{}".format(row +1)]
    if "<MergedCell 'Sheet'" in str(row_status):
        continue
    row_status.value = i
    i += 1

ws["b{}".format(ws.max_row)].border = Border(bottom=Side(style="thick"), left=Side(style="thin"), right=Side(style="thin"))
ws["A{}".format(ws.max_row)].border = Border(bottom=Side(style="thick"), left=Side(style="thick"), right=Side(style="thin"))

ws.freeze_panes = "B2"
wb.save("./exercise/exelpracticefiles/Programing.xlsx")