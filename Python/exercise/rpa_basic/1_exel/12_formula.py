import datetime
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws["A1"] = datetime.datetime.today()
ws["A2"] = "=Sum(1, 2, 3)"
ws["A3"] = "=AVERAGE(1, 2, 3)"

ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "=SUM(A4:A5)"

ws.column_dimensions["A"].width =20
wb.save("sample_formula.xlsx")