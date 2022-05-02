from openpyxl.styles import * # Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]

# A 열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5
# 1 행의 높이를 50으로 설정
ws.row_dimensions[1].height = 50

a1.font = Font(color="FF0000", italic=True, bold=True) # 글자 색은 빨갛게, 이탤릭, 두껍게
b1.font = Font(color="CC33FF", name="Arial", strike=True) # 폰트를 Arial 로 설정, 
c1.font = Font(color="0000FF", size=20, underline="single") # 글자 크기를 20으로, 밑줄 적용

# 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), bottom=Side(style="thin"), top=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90 점 넘는 셀에 대해서 초록색으로 적용
for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center") # 각 cell에 대해 가운데 정렬
        if cell.column == 1: # A 번호열은 제외
            continue
        
        # cell이 정수형 데이터이고 90점보다 높으면
        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid") # 배경을 초록색으로 설정
            cell.font = Font(color="FF0000") # 폰트 색상 빨간색

ws.freeze_panes = "B2" # B2 기준으로 고정
# wb.close()
wb.save("sample_style.xlsx")